import csv
import datetime
import difflib
import json
import re
import shutil
import subprocess
from pathlib import Path

import openpyxl
from agents import function_tool


@function_tool
def read_excel(
    file_path: str,
    sheet_name: str = "",
    start_row: int = 1,
    max_rows: int = 50,
) -> str:
    """Excelファイルを読み込み、ヘッダーと行データを構造化データとして返す。

    Args:
        file_path: Excelファイルの絶対パスまたは相対パス。
        sheet_name: 読み込むシート名。空文字列の場合はアクティブシートを使用する。
        start_row: 読み込み開始データ行番号（1始まり、ヘッダー行の次を1とする）。
        max_rows: 最大読み込み行数（1〜100）。大きなファイルではページネーションに使用する。
    """
    try:
        p = Path(file_path)
        if not p.exists():
            return json.dumps(
                {"error": f"ファイルが見つかりません: {file_path}"},
                ensure_ascii=False,
            )

        wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            if ws is None:
                return json.dumps(
                    {"error": "アクティブなシートがありません"},
                    ensure_ascii=False,
                )

            # ヘッダー行（1行目）を読み取る
            headers: list[str] = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else "")

            if not any(headers):
                return json.dumps(
                    {"error": "シートにデータがありません"},
                    ensure_ascii=False,
                )

            # データ行数を算出
            total_data_rows = ws.max_row - 1 if ws.max_row else 0

            # ページネーション
            max_rows = max(1, min(100, max_rows))
            actual_start = start_row + 1  # Excel行番号（ヘッダーが1行目）
            actual_end = min(actual_start + max_rows - 1, ws.max_row)

            rows = []
            for row in ws.iter_rows(
                min_row=actual_start, max_row=actual_end, values_only=False
            ):
                row_data: dict = {"excel_row": row[0].row}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        key = headers[i] if headers[i] else f"列{i + 1}"
                        row_data[key] = cell.value
                rows.append(row_data)

            returned_count = len(rows)
            return json.dumps(
                {
                    "file": str(p.resolve()),
                    "sheet": ws.title,
                    "headers": headers,
                    "rows": rows,
                    "total_data_rows": total_data_rows,
                    "returned_range": f"{start_row}-{start_row + returned_count - 1} / {total_data_rows}",
                },
                ensure_ascii=False,
                default=str,
            )
        finally:
            wb.close()

    except Exception as e:
        return json.dumps(
            {"error": f"Excelファイルの読み込みに失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def read_source_code(
    file_path: str,
    start_line: int = 1,
    end_line: int = 0,
) -> str:
    """ファイルを読み込み、指定行範囲の内容を行番号付きで返す。

    Args:
        file_path: ファイルの絶対パスまたは相対パス。
        start_line: 読み込み開始行番号（1始まり）。
        end_line: 読み込み終了行番号。0の場合はstart_lineから50行分を読み込む。
    """
    try:
        p = Path(file_path)
        if not p.exists():
            return json.dumps(
                {"error": f"ファイルが見つかりません: {file_path}"},
                ensure_ascii=False,
            )
        if not p.is_file():
            return json.dumps(
                {"error": f"ファイルではありません: {file_path}"},
                ensure_ascii=False,
            )

        # エンコーディング自動検出（日本語対応）
        content = None
        used_encoding = None
        for enc in ("utf-8", "utf-8-sig", "shift_jis", "cp932", "latin-1"):
            try:
                content = p.read_text(encoding=enc)
                used_encoding = enc
                break
            except (UnicodeDecodeError, UnicodeError):
                continue

        if content is None:
            return json.dumps(
                {"error": f"ファイルのエンコーディングを検出できません: {file_path}"},
                ensure_ascii=False,
            )

        lines = content.splitlines()
        total_lines = len(lines)

        if end_line <= 0:
            end_line = start_line + 49

        start_line = max(1, start_line)
        end_line = min(end_line, total_lines)

        selected = lines[start_line - 1 : end_line]
        numbered = "\n".join(
            f"{start_line + i}: {line}" for i, line in enumerate(selected)
        )

        return json.dumps(
            {
                "file": str(p.resolve()),
                "start_line": start_line,
                "end_line": end_line,
                "total_lines": total_lines,
                "encoding": used_encoding,
                "content": numbered,
            },
            ensure_ascii=False,
        )

    except PermissionError:
        return json.dumps(
            {"error": f"ファイルの読み込み権限がありません: {file_path}"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"ファイルの読み込みに失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def write_excel_cells(
    file_path: str,
    updates_json: str,
    sheet_name: str = "",
) -> str:
    """Excelファイルの指定セルに値を一括書き込みし、ファイルを保存する。

    Args:
        file_path: Excelファイルの絶対パスまたは相対パス。
        updates_json: 書き込むセル情報のJSON配列。形式は [{"row": Excel行番号, "column": "列名", "value": "値"}, ...] とする。
        sheet_name: 書き込むシート名。空文字列の場合はアクティブシートを使用する。
    """
    try:
        p = Path(file_path)
        if not p.exists():
            return json.dumps(
                {"error": f"ファイルが見つかりません: {file_path}"},
                ensure_ascii=False,
            )

        try:
            updates = json.loads(updates_json)
        except json.JSONDecodeError as e:
            return json.dumps(
                {"error": f"updates_jsonのJSON形式が不正です: {e}"},
                ensure_ascii=False,
            )

        if not isinstance(updates, list):
            return json.dumps(
                {"error": "updates_jsonはJSON配列でなければなりません"},
                ensure_ascii=False,
            )

        wb = openpyxl.load_workbook(str(p))
        ws = wb[sheet_name] if sheet_name else wb.active
        if ws is None:
            wb.close()
            return json.dumps(
                {"error": "アクティブなシートがありません"},
                ensure_ascii=False,
            )

        # ヘッダー名 → 列番号のマッピングを構築
        header_map: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val is not None:
                header_map[str(val)] = col_idx

        updated_count = 0
        for update in updates:
            row = update.get("row")
            column_name = update.get("column")
            value = update.get("value")

            if row is None or column_name is None:
                continue

            # 列名が存在しなければ新規列を作成
            if column_name not in header_map:
                new_col = ws.max_column + 1
                ws.cell(row=1, column=new_col, value=column_name)
                header_map[column_name] = new_col

            col_idx = header_map[column_name]
            ws.cell(row=row, column=col_idx, value=value)
            updated_count += 1

        wb.save(str(p))
        wb.close()

        return json.dumps(
            {
                "status": "success",
                "updated_cells": updated_count,
                "file": str(p.resolve()),
            },
            ensure_ascii=False,
        )

    except PermissionError:
        return json.dumps(
            {
                "error": f"ファイルが他のプログラムで開かれている可能性があります。Excelを閉じてから再試行してください: {file_path}"
            },
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"Excelファイルの書き込みに失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def list_directory(
    directory_path: str,
    pattern: str = "*",
    recursive: bool = False,
) -> str:
    """ディレクトリ内のファイルとサブディレクトリの一覧を返す。

    Args:
        directory_path: 一覧を取得するディレクトリの絶対パスまたは相対パス。
        pattern: ファイル名フィルタ（globパターン。例: '*.c', '*.py'）。
        recursive: Trueの場合サブディレクトリも再帰的に検索する。パターン例: '*.py' で全階層の.pyファイルを検索できる。
    """
    try:
        p = Path(directory_path)
        if not p.exists():
            return json.dumps(
                {"error": f"ディレクトリが見つかりません: {directory_path}"},
                ensure_ascii=False,
            )
        if not p.is_dir():
            return json.dumps(
                {"error": f"ディレクトリではありません: {directory_path}"},
                ensure_ascii=False,
            )

        if recursive:
            glob_pattern = f"**/{pattern}"
        else:
            glob_pattern = pattern

        entries = sorted(p.glob(glob_pattern))
        dirs = []
        files = []
        limit = 200

        for entry in entries:
            if len(dirs) + len(files) >= limit:
                break
            if entry.is_dir():
                # 再帰時は相対パスを表示
                rel = entry.relative_to(p)
                dirs.append(str(rel) + "/")
            elif entry.is_file():
                rel = entry.relative_to(p)
                files.append(str(rel))

        total = len(dirs) + len(files)
        total_found = len(entries)
        return json.dumps(
            {
                "directory": str(p.resolve()),
                "pattern": pattern,
                "recursive": recursive,
                "directories": dirs,
                "files": files,
                "total_entries": total,
                "truncated": total_found > limit,
                "truncated_note": f"表示は{limit}件までです（全{total_found}件）" if total_found > limit else None,
            },
            ensure_ascii=False,
        )

    except PermissionError:
        return json.dumps(
            {"error": f"ディレクトリの読み込み権限がありません: {directory_path}"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"ディレクトリの一覧取得に失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def write_file(
    file_path: str,
    content: str,
    encoding: str = "utf-8",
) -> str:
    """テキストファイルを新規作成または上書き保存する。

    Args:
        file_path: 保存先ファイルの絶対パスまたは相対パス。
        content: ファイルに書き込む内容。
        encoding: 文字エンコーディング。デフォルトは 'utf-8'。日本語Windowsでは 'shift_jis' や 'cp932' も指定可能。
    """
    try:
        p = Path(file_path)

        # 親ディレクトリが存在しない場合は作成
        p.parent.mkdir(parents=True, exist_ok=True)

        p.write_text(content, encoding=encoding)

        return json.dumps(
            {
                "status": "success",
                "file": str(p.resolve()),
                "size_bytes": p.stat().st_size,
                "encoding": encoding,
            },
            ensure_ascii=False,
        )

    except PermissionError:
        return json.dumps(
            {"error": f"ファイルの書き込み権限がありません: {file_path}"},
            ensure_ascii=False,
        )
    except LookupError:
        return json.dumps(
            {"error": f"不明なエンコーディングです: {encoding}（utf-8, shift_jis, cp932 などを指定してください）"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"ファイルの書き込みに失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def search_in_file(
    path: str,
    pattern: str,
    is_regex: bool = False,
    recursive: bool = False,
    file_pattern: str = "*",
    max_results: int = 50,
) -> str:
    """ファイル内またはディレクトリ内でテキストを検索し、マッチした行を返す。

    Args:
        path: 検索対象のファイルパスまたはディレクトリパス。
        pattern: 検索する文字列または正規表現パターン。
        is_regex: Trueの場合、patternを正規表現として扱う。
        recursive: ディレクトリ検索時にサブディレクトリも再帰的に検索する。
        file_pattern: ディレクトリ検索時のファイル名フィルタ（globパターン。例: '*.c', '*.py'）。
        max_results: 最大結果件数（1〜200）。
    """
    try:
        p = Path(path)
        if not p.exists():
            return json.dumps(
                {"error": f"パスが見つかりません: {path}"},
                ensure_ascii=False,
            )

        max_results = max(1, min(200, max_results))

        # 正規表現のコンパイル
        if is_regex:
            try:
                regex = re.compile(pattern)
            except re.error as e:
                return json.dumps(
                    {"error": f"正規表現が不正です: {e}"},
                    ensure_ascii=False,
                )
        else:
            regex = None

        def search_file(file_path: Path) -> list[dict]:
            """単一ファイルを検索してマッチ行を返す。"""
            results = []
            content = None
            for enc in ("utf-8", "utf-8-sig", "shift_jis", "cp932", "latin-1"):
                try:
                    content = file_path.read_text(encoding=enc)
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            if content is None:
                return results

            for line_num, line in enumerate(content.splitlines(), 1):
                matched = False
                if regex:
                    matched = bool(regex.search(line))
                else:
                    matched = pattern in line
                if matched:
                    results.append({
                        "file": str(file_path.resolve()),
                        "line": line_num,
                        "content": line.rstrip(),
                    })
            return results

        matches: list[dict] = []

        if p.is_file():
            matches = search_file(p)
        elif p.is_dir():
            glob_pat = f"**/{file_pattern}" if recursive else file_pattern
            for entry in sorted(p.glob(glob_pat)):
                if entry.is_file():
                    matches.extend(search_file(entry))
                    if len(matches) >= max_results:
                        break
        else:
            return json.dumps(
                {"error": f"ファイルでもディレクトリでもありません: {path}"},
                ensure_ascii=False,
            )

        truncated = len(matches) > max_results
        matches = matches[:max_results]

        return json.dumps(
            {
                "pattern": pattern,
                "is_regex": is_regex,
                "matches": matches,
                "total_matches": len(matches),
                "truncated": truncated,
            },
            ensure_ascii=False,
        )

    except PermissionError:
        return json.dumps(
            {"error": f"読み込み権限がありません: {path}"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"検索に失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def get_file_info(file_path: str) -> str:
    """ファイルのメタデータ（サイズ、更新日時、エンコーディング推定など）を取得する。

    Args:
        file_path: 情報を取得するファイルの絶対パスまたは相対パス。
    """
    try:
        p = Path(file_path)
        if not p.exists():
            return json.dumps(
                {"error": f"ファイルが見つかりません: {file_path}"},
                ensure_ascii=False,
            )

        stat = p.stat()
        size_bytes = stat.st_size

        # 人間が読みやすいサイズ表記
        if size_bytes < 1024:
            size_human = f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            size_human = f"{size_bytes / 1024:.1f} KB"
        elif size_bytes < 1024 * 1024 * 1024:
            size_human = f"{size_bytes / (1024 * 1024):.1f} MB"
        else:
            size_human = f"{size_bytes / (1024 * 1024 * 1024):.1f} GB"

        modified = datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        created = datetime.datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S")

        # エンコーディング推定（テキストファイルの場合）
        detected_encoding = None
        if p.is_file():
            for enc in ("utf-8", "utf-8-sig", "shift_jis", "cp932"):
                try:
                    p.read_text(encoding=enc)
                    detected_encoding = enc
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
                except Exception:
                    break

        result = {
            "file": str(p.resolve()),
            "name": p.name,
            "extension": p.suffix,
            "is_file": p.is_file(),
            "is_directory": p.is_dir(),
            "size_bytes": size_bytes,
            "size_human": size_human,
            "modified": modified,
            "created": created,
        }
        if detected_encoding:
            result["estimated_encoding"] = detected_encoding

        return json.dumps(result, ensure_ascii=False)

    except PermissionError:
        return json.dumps(
            {"error": f"ファイルの情報取得権限がありません: {file_path}"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"ファイル情報の取得に失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def read_excel_sheet_names(file_path: str) -> str:
    """Excelファイルに含まれる全シート名の一覧を取得する。

    Args:
        file_path: Excelファイルの絶対パスまたは相対パス。
    """
    try:
        p = Path(file_path)
        if not p.exists():
            return json.dumps(
                {"error": f"ファイルが見つかりません: {file_path}"},
                ensure_ascii=False,
            )

        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        try:
            sheet_names = wb.sheetnames
            active_sheet = wb.active.title if wb.active else None

            return json.dumps(
                {
                    "file": str(p.resolve()),
                    "sheet_names": sheet_names,
                    "sheet_count": len(sheet_names),
                    "active_sheet": active_sheet,
                },
                ensure_ascii=False,
            )
        finally:
            wb.close()

    except Exception as e:
        return json.dumps(
            {"error": f"Excelファイルのシート一覧取得に失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def copy_file(
    source_path: str,
    destination_path: str,
    operation: str = "copy",
) -> str:
    """ファイルまたはディレクトリをコピーまたは移動する。

    Args:
        source_path: コピー/移動元のファイルまたはディレクトリのパス。
        destination_path: コピー/移動先のパス。
        operation: 操作の種類。'copy' でコピー、'move' で移動。
    """
    try:
        src = Path(source_path)
        dst = Path(destination_path)

        if not src.exists():
            return json.dumps(
                {"error": f"コピー元が見つかりません: {source_path}"},
                ensure_ascii=False,
            )

        if operation not in ("copy", "move"):
            return json.dumps(
                {"error": f"operationは 'copy' または 'move' を指定してください（指定値: {operation}）"},
                ensure_ascii=False,
            )

        # 移動先の親ディレクトリを作成
        dst.parent.mkdir(parents=True, exist_ok=True)

        if operation == "copy":
            if src.is_dir():
                shutil.copytree(str(src), str(dst))
            else:
                shutil.copy2(str(src), str(dst))
            action = "コピー"
        else:
            shutil.move(str(src), str(dst))
            action = "移動"

        return json.dumps(
            {
                "status": "success",
                "operation": action,
                "source": str(src.resolve()),
                "destination": str(Path(destination_path).resolve()),
            },
            ensure_ascii=False,
        )

    except PermissionError:
        return json.dumps(
            {"error": f"ファイル操作の権限がありません: {source_path} → {destination_path}"},
            ensure_ascii=False,
        )
    except shutil.Error as e:
        return json.dumps(
            {"error": f"ファイル操作に失敗しました: {e}"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"ファイル操作に失敗しました: {e}"},
            ensure_ascii=False,
        )


# run_command の許可コマンドリスト
_ALLOWED_COMMANDS = {
    # ファイル操作（読み取り系）
    "dir", "type", "findstr", "tree", "where",
    # ビルドツール
    "python", "pip", "gcc", "g++", "make", "cmake",
    # 静的解析ツール
    "cppcheck", "clang-tidy", "flake8", "pylint", "mypy",
    # バージョン管理（読み取り系）
    "git",
}

# シェルメタ文字（コマンドインジェクション防止）
_SHELL_METACHAR_PATTERN = re.compile(r"[&|;><`$]")


@function_tool
def run_command(
    command: str,
    working_directory: str = "",
    timeout: int = 30,
) -> str:
    """ホワイトリストに登録されたコマンドを安全に実行する。

    Args:
        command: 実行するコマンド文字列（例: 'python --version', 'gcc -Wall main.c'）。
        working_directory: コマンドの作業ディレクトリ。空文字列の場合はカレントディレクトリを使用する。
        timeout: タイムアウト秒数（1〜120）。デフォルトは30秒。
    """
    try:
        if not command.strip():
            return json.dumps(
                {"error": "コマンドが空です"},
                ensure_ascii=False,
            )

        # シェルメタ文字チェック
        if _SHELL_METACHAR_PATTERN.search(command):
            return json.dumps(
                {"error": "セキュリティ上の理由により、シェルメタ文字（&, |, ;, >, <, `, $）を含むコマンドは実行できません"},
                ensure_ascii=False,
            )

        # コマンドを分割して先頭のプログラム名を取得
        parts = command.split()
        program = parts[0].lower()

        # パスからプログラム名だけを抽出
        program_name = Path(program).stem

        if program_name not in _ALLOWED_COMMANDS:
            return json.dumps(
                {
                    "error": f"許可されていないコマンドです: {program_name}",
                    "allowed_commands": sorted(_ALLOWED_COMMANDS),
                },
                ensure_ascii=False,
            )

        # タイムアウト制限
        timeout = max(1, min(120, timeout))

        # 作業ディレクトリの確認
        cwd = None
        if working_directory:
            wd = Path(working_directory)
            if not wd.exists() or not wd.is_dir():
                return json.dumps(
                    {"error": f"作業ディレクトリが見つかりません: {working_directory}"},
                    ensure_ascii=False,
                )
            cwd = str(wd)

        result = subprocess.run(
            parts,
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=cwd,
            shell=False,
        )

        output = {
            "command": command,
            "return_code": result.returncode,
        }
        if result.stdout:
            # 出力が長すぎる場合は切り詰め
            stdout = result.stdout
            if len(stdout) > 10000:
                stdout = stdout[:10000] + f"\n... (出力を切り詰めました。全{len(result.stdout)}文字)"
            output["stdout"] = stdout
        if result.stderr:
            stderr = result.stderr
            if len(stderr) > 5000:
                stderr = stderr[:5000] + f"\n... (エラー出力を切り詰めました。全{len(result.stderr)}文字)"
            output["stderr"] = stderr

        return json.dumps(output, ensure_ascii=False)

    except subprocess.TimeoutExpired:
        return json.dumps(
            {"error": f"コマンドがタイムアウトしました（{timeout}秒）: {command}"},
            ensure_ascii=False,
        )
    except FileNotFoundError:
        return json.dumps(
            {"error": f"コマンドが見つかりません: {parts[0]}（パスが通っているか確認してください）"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"コマンドの実行に失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def diff_files(
    file_path_a: str,
    file_path_b: str,
    context_lines: int = 3,
) -> str:
    """2つのファイルを比較し、差分をunified diff形式で返す。

    Args:
        file_path_a: 比較元ファイルのパス。
        file_path_b: 比較先ファイルのパス。
        context_lines: 差分の前後に表示するコンテキスト行数（0〜10）。
    """
    try:
        pa = Path(file_path_a)
        pb = Path(file_path_b)

        if not pa.exists():
            return json.dumps(
                {"error": f"ファイルが見つかりません: {file_path_a}"},
                ensure_ascii=False,
            )
        if not pb.exists():
            return json.dumps(
                {"error": f"ファイルが見つかりません: {file_path_b}"},
                ensure_ascii=False,
            )

        context_lines = max(0, min(10, context_lines))

        # ファイル読み込み（エンコーディング自動検出）
        def read_with_detect(fp: Path) -> list[str] | None:
            for enc in ("utf-8", "utf-8-sig", "shift_jis", "cp932", "latin-1"):
                try:
                    return fp.read_text(encoding=enc).splitlines(keepends=True)
                except (UnicodeDecodeError, UnicodeError):
                    continue
            return None

        lines_a = read_with_detect(pa)
        if lines_a is None:
            return json.dumps(
                {"error": f"ファイルのエンコーディングを検出できません: {file_path_a}"},
                ensure_ascii=False,
            )

        lines_b = read_with_detect(pb)
        if lines_b is None:
            return json.dumps(
                {"error": f"ファイルのエンコーディングを検出できません: {file_path_b}"},
                ensure_ascii=False,
            )

        diff = list(difflib.unified_diff(
            lines_a,
            lines_b,
            fromfile=str(pa),
            tofile=str(pb),
            n=context_lines,
        ))

        if not diff:
            return json.dumps(
                {
                    "file_a": str(pa.resolve()),
                    "file_b": str(pb.resolve()),
                    "has_diff": False,
                    "message": "ファイルは同一です",
                },
                ensure_ascii=False,
            )

        # 変更行数の集計
        added = sum(1 for line in diff if line.startswith("+") and not line.startswith("+++"))
        removed = sum(1 for line in diff if line.startswith("-") and not line.startswith("---"))

        diff_text = "".join(diff)
        if len(diff_text) > 10000:
            diff_text = diff_text[:10000] + f"\n... (差分を切り詰めました)"

        return json.dumps(
            {
                "file_a": str(pa.resolve()),
                "file_b": str(pb.resolve()),
                "has_diff": True,
                "added_lines": added,
                "removed_lines": removed,
                "diff": diff_text,
            },
            ensure_ascii=False,
        )

    except PermissionError:
        return json.dumps(
            {"error": "ファイルの読み込み権限がありません"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"ファイル比較に失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def append_to_file(
    file_path: str,
    content: str,
    encoding: str = "utf-8",
) -> str:
    """既存ファイルの末尾に内容を追記する。ファイルが存在しない場合は新規作成する。

    Args:
        file_path: 追記先ファイルの絶対パスまたは相対パス。
        content: 追記する内容。
        encoding: 文字エンコーディング。デフォルトは 'utf-8'。日本語Windowsでは 'shift_jis' や 'cp932' も指定可能。
    """
    try:
        p = Path(file_path)
        p.parent.mkdir(parents=True, exist_ok=True)

        with open(str(p), "a", encoding=encoding) as f:
            f.write(content)

        return json.dumps(
            {
                "status": "success",
                "file": str(p.resolve()),
                "size_bytes": p.stat().st_size,
                "encoding": encoding,
            },
            ensure_ascii=False,
        )

    except PermissionError:
        return json.dumps(
            {"error": f"ファイルの書き込み権限がありません: {file_path}"},
            ensure_ascii=False,
        )
    except LookupError:
        return json.dumps(
            {"error": f"不明なエンコーディングです: {encoding}（utf-8, shift_jis, cp932 などを指定してください）"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"ファイルへの追記に失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def create_excel_sheet(
    file_path: str,
    sheet_name: str,
    position: int = -1,
) -> str:
    """既存のExcelファイルに新しいシートを追加する。

    Args:
        file_path: Excelファイルの絶対パスまたは相対パス。
        sheet_name: 新しいシートの名前。
        position: シートの挿入位置（0始まりのインデックス）。-1の場合は末尾に追加する。
    """
    try:
        p = Path(file_path)
        if not p.exists():
            return json.dumps(
                {"error": f"ファイルが見つかりません: {file_path}"},
                ensure_ascii=False,
            )

        wb = openpyxl.load_workbook(str(p))
        try:
            if sheet_name in wb.sheetnames:
                wb.close()
                return json.dumps(
                    {"error": f"シート '{sheet_name}' は既に存在します"},
                    ensure_ascii=False,
                )

            if position < 0:
                position = len(wb.sheetnames)

            wb.create_sheet(title=sheet_name, index=position)
            wb.save(str(p))

            return json.dumps(
                {
                    "status": "success",
                    "file": str(p.resolve()),
                    "created_sheet": sheet_name,
                    "all_sheets": wb.sheetnames,
                },
                ensure_ascii=False,
            )
        finally:
            wb.close()

    except PermissionError:
        return json.dumps(
            {
                "error": f"ファイルが他のプログラムで開かれている可能性があります。Excelを閉じてから再試行してください: {file_path}"
            },
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"シートの作成に失敗しました: {e}"},
            ensure_ascii=False,
        )


@function_tool
def export_excel_to_csv(
    file_path: str,
    output_path: str,
    sheet_name: str = "",
    encoding: str = "utf-8",
) -> str:
    """ExcelファイルのシートをCSVファイルとしてエクスポートする。

    Args:
        file_path: Excelファイルの絶対パスまたは相対パス。
        output_path: 出力先CSVファイルのパス。
        sheet_name: エクスポートするシート名。空文字列の場合はアクティブシートを使用する。
        encoding: CSV出力時の文字エンコーディング。デフォルトは 'utf-8'。
    """
    try:
        p = Path(file_path)
        if not p.exists():
            return json.dumps(
                {"error": f"ファイルが見つかりません: {file_path}"},
                ensure_ascii=False,
            )

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            if ws is None:
                return json.dumps(
                    {"error": "アクティブなシートがありません"},
                    ensure_ascii=False,
                )

            row_count = 0
            with open(str(out), "w", encoding=encoding, newline="") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(
                        [str(cell) if cell is not None else "" for cell in row]
                    )
                    row_count += 1

            return json.dumps(
                {
                    "status": "success",
                    "source_file": str(p.resolve()),
                    "source_sheet": ws.title,
                    "output_file": str(out.resolve()),
                    "row_count": row_count,
                    "encoding": encoding,
                },
                ensure_ascii=False,
            )
        finally:
            wb.close()

    except PermissionError:
        return json.dumps(
            {"error": f"ファイル操作の権限がありません"},
            ensure_ascii=False,
        )
    except LookupError:
        return json.dumps(
            {"error": f"不明なエンコーディングです: {encoding}"},
            ensure_ascii=False,
        )
    except Exception as e:
        return json.dumps(
            {"error": f"CSVエクスポートに失敗しました: {e}"},
            ensure_ascii=False,
        )
